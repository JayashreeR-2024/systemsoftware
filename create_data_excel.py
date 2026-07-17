from faker import Faker
from openpyxl import Workbook

fake = Faker()


# Create workbook
workbook = Workbook()

# Create sheet
sheet = workbook.active
sheet.title = "Login"


# Add headers
sheet["A1"] = "username"
sheet["B1"] = "password"
sheet["C1"] = "email"
sheet["D1"] = "first_name"
sheet["E1"] = "last_name"


# Generate fake data
for row in range(2, 11):

    sheet.cell(
        row=row,
        column=1
    ).value = fake.user_name()


    sheet.cell(
        row=row,
        column=2
    ).value = fake.password(
        length=10
    )


    sheet.cell(
        row=row,
        column=3
    ).value = fake.email()


    sheet.cell(
        row=row,
        column=4
    ).value = fake.first_name()


    sheet.cell(
        row=row,
        column=5
    ).value = fake.last_name()



# Save Excel file
workbook.save("testdata.xlsx")


print("✅ testdata.xlsx created successfully")