from faker import Faker
from openpyxl import Workbook

fake = Faker()

for file_num in range(1, 6):
    wb = Workbook()
    ws = wb.active
    ws.title = "Fake Data"

    ws.cell(row=1, column=1).value = "Name"
    ws.cell(row=1, column=2).value = "Email"

    for row in range(2, 12):
        ws.cell(row=row, column=1).value = fake.name()
        ws.cell(row=row, column=2).value = fake.email()


    wb.save(f"fake_data_{file_num}.xlsx")

print("5 Excel workbooks created successfully!")