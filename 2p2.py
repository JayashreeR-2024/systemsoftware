from faker import Faker
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Fake Data"

fake_data = Faker()

ws.cell(row=1, column=1).value = "Name"
ws.cell(row=1, column=2).value = "Email"

for i in range(2, 12):
    ws.cell(row=i, column=1).value = fake_data.name()
    ws.cell(row=i, column=2).value = fake_data.email()

wb.save("fake_data.xlsx")

print("Excel file 'fake_data.xlsx' created successfully!")